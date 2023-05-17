# ///////////////////////////////////////////////////////////////
#
# BY: WANDERSON M.PIMENTA
# PROJECT MADE WITH: Qt Designer and PySide6
# V: 1.0.0
#
# This project can be used freely for all uses, as long as they maintain the
# respective credits only in the Python scripts, any information in the visual
# interface (GUI) can be modified without any implication.
#
# There are limitations on Qt licenses if you want to use your products
# commercially, I recommend reading them on the official website:
# https://doc.qt.io/qtforpython/licenses.html
#
# ///////////////////////////////////////////////////////////////

import sys
import os
import platform
from functools import partial

# IMPORT / GUI AND MODULES AND WIDGETS
# ///////////////////////////////////////////////////////////////
from modules import *
from modules.ui_main import MyGLViewWidget
from widgets import *
import json
from xls2xlsx import XLS2XLSX
import openpyxl
from openpyxl.styles import PatternFill, Color

from PySide6 import QtWidgets, QtCore
from PySide6.QtCore import Qt
os.environ["QT_FONT_DPI"] = "96" # FIX Problem for High DPI and Scale above 100%

# SET AS GLOBAL WIDGETS
# ///////////////////////////////////////////////////////////////
widgets = None

class NumericItem(QtWidgets.QTableWidgetItem):
    def __lt__(self, other):
        return ((self.data(QtCore.Qt.UserRole)) <
                (other.data(QtCore.Qt.UserRole)))
class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        # widgets.new_page.setMouseTracking(True)
        # self.setMouseTracking(True)
        # self.installEventFilter(self)
        # SET AS GLOBAL WIDGETS
        # ///////////////////////////////////////////////////////////////
        self.is_authorized = False
        self.accounts = {"admin": "admin1234"}

        self.lines = []
        self.pen = QPen(Qt.black)
        self.min_orange = QColor(51, 33, 0).getRgbF()[:3]
        self.max_orange = QColor(255, 165, 0).getRgbF()[:3]
        self.min_red = QColor(51, 0, 0).getRgbF()[:3]
        self.max_red = QColor(255, 165, 0).getRgbF()[:3]
        self.min_yellow = QColor(51, 51, 0).getRgbF()[:3]
        self.max_yellow = QColor(255, 255, 0).getRgbF()[:3]
        self.min_green = QColor(0, 51, 0).getRgbF()[:3]
        self.max_green = QColor(0, 255, 0).getRgbF()[:3]
        self.min_white = QColor(51, 51, 51).getRgbF()[:3]
        self.max_white = QColor(255, 255, 255).getRgbF()[:3]
        self.price_column = 4
        self.factory_column = 5
        self.mouse_position = None
        self.a_descr_column = 1
        file = open('variables.json')
        self.data = json.load(file)
        print(self.data["file"])
        self.tempaddr = self.data["file"]  # "G:\\FIles\\Products\\Product Test.xlsx"
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.flag_generate_factorys = True

        global widgets
        widgets = self.ui
        for col in range(5):
            item = widgets.tableWidget.item(0, col)
            item.setFlags(item.flags() & ~Qt.ItemIsSelectable)
        widgets.btn_widgets.setVisible(False)
        widgets.btn_new.setVisible(False)
        self.keys = []
        self.values = []
        self.a_number_column = 0
        self.havetb_column = 2
        self.full_costs = {"global":0}
        self.stock_column = 3
        self.forecolors = {}
        self.dictionary = dict()
        self.dictionary_all_values = dict()
        self.dict_values_factory = []
        self.firstTable = False
        self.secondTable = False
        self.current_page = "home"
        self.set_factorys = set()
        self.colors_chart = [
            (1, 0, 0, 55),
            (1, 0.65, 0, 255),
            (1, 1, 0, 255),
            (0, 1, 0, 255),
            (1, 1, 1, 255)
        ]
        self.gl_view = MyGLViewWidget()
        self.gl_view.setCameraPosition(distance=5, elevation=30, azimuth=40)
        self.gl_view.setBackgroundColor((0, 0, 0, 0))
        self.buttons_red = {}
        self.data_costs = {"global":[0,0,0,0,0]}
        self.buttons_orange = {}
        self.buttons_yellow = {}
        self.factory_keys = []
        # self.detector = ColorDetector()
        self.setMouseTracking(True)  # Enable mouse tracking
        # self.detector.show()
        # USE CUSTOM TITLE BAR | USE AS "False" FOR MAC OR LINUX
        # ///////////////////////////////////////////////////////////////
        Settings.ENABLE_CUSTOM_TITLE_BAR = True
        self.Must_have()
        # APP NAME
        # ///////////////////////////////////////////////////////////////
        title = "PyDracula - Modern GUI"
        description = "Westa GmbH Product control"
        # APPLY TEXTS
        self.setWindowTitle(title)
        widgets.titleRightInfo.setText(description)
        self.end_non_empty = -1
        self.dictionary_factorys = {}
        self.keys_for_dict = []
        # TOGGLE MENU
        # ///////////////////////////////////////////////////////////////
        widgets.toggleButton.clicked.connect(lambda: UIFunctions.toggleMenu(self, True))

        # SET UI DEFINITIONS
        # ///////////////////////////////////////////////////////////////
        UIFunctions.uiDefinitions(self)
        widgets.toggleLeftBox.clicked.connect(lambda: self.openCloseLeftBox(self.current_page))
        widgets.extraCloseColumnBtn.clicked.connect(lambda: self.openCloseLeftBox(self.current_page))
        # QTableWidget PARAMETERS
        # ///////////////////////////////////////////////////////////////
        widgets.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        widgets.tableWidget.cellClicked.connect(self.CellClicked)
        widgets.lineEdit.textChanged.connect(self.search_table)
        # BUTTONS CLICK
        # ///////////////////////////////////////////////////////////////
        # LEFT MENUS
        widgets.btn_home.clicked.connect(self.buttonClick)
        widgets.btn_widgets.clicked.connect(self.buttonClick)
        widgets.btn_new.clicked.connect(self.buttonClick)
        widgets.pushButton.clicked.connect(self.buttonClick)
        widgets.btn_export.clicked.connect(self.buttonClick)
        widgets.button_select.clicked.connect(self.buttonClick)
        widgets.button1.clicked.connect(self.buttonClick)
        widgets.button2.clicked.connect(self.buttonClick)
        widgets.button3.clicked.connect(self.buttonClick)
        widgets.button4.clicked.connect(self.buttonClick)
        widgets.button5.clicked.connect(self.buttonClick)
        widgets.button6.clicked.connect(self.buttonClick)
        widgets.button7.clicked.connect(self.buttonClick)
        widgets.button8.clicked.connect(self.buttonClick)
        widgets.button9.clicked.connect(self.buttonClick)
        widgets.button10.clicked.connect(self.buttonClick)
        widgets.btn_global.clicked.connect(self.buttonClick)
        widgets.button_reset.clicked.connect(self.buttonClick)
        widgets.login_button.clicked.connect(self.buttonClick)
        self.buttonarray = [widgets.button1, widgets.button2, widgets.button3, widgets.button4,
                         widgets.button5, widgets.button6, widgets.button7, widgets.button8, widgets.button9, widgets.button10]
        self.sorting_design(widgets.button8)
        self.already_created = False
        self.labelHello =  QLabel('Hello World', widgets.new_page)

        self.labelHello.setWordWrap(True)
        self.labelHello.setMinimumHeight(60)
        self.labelHello.move(100,100)
        legend_list = QListWidget(widgets.new_page)
        legend_list.setStyleSheet("background-color: #252930; border: none;")
        legend_list.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        # Create a dictionary mapping colors to labels
        label_colors = {
            "red": "Difference > 100",
            "orange": "Difference > 50",
            "yellow": "Difference > 0",
            "green": "In stock > have to be",
            "white": "In stock >> have to be"
        }

        # Add items to the legend list with labels and colors
        for color_name, label_text in label_colors.items():
            item = QListWidgetItem(legend_list)
            item_widget = QWidget()
            item_layout = QVBoxLayout(item_widget)

            color_label = QLabel(widgets.new_page)
            color_label.setFixedWidth(20)
            color = QColor(color_name)
            color_label.setStyleSheet(f"background-color: {color.name()};")

            text_label = QLabel(label_text, widgets.new_page)
            text_label.setFont(QFont("Arial", 10))
            Qw = QWidget()
            ql = QHBoxLayout(Qw)
            ql.addWidget(color_label)
            ql.addWidget(text_label)
            Qw.setLayout(ql)
            item_layout.addWidget(Qw)

            item_layout.setContentsMargins(0, 0, 0, 0)
            item_widget.setLayout(item_layout)

            item.setSizeHint(item_widget.sizeHint())
            legend_list.addItem(item)
            legend_list.setItemWidget(item, item_widget)

        # Set the legend list to align top-right


        self.another_dict = {}
        # EXTRA LEFT BOX


        # EXTRA RIGHT BOX
        def openCloseRightBox():
            UIFunctions.toggleRightBox(self, True)
        widgets.settingsTopBtn.clicked.connect(openCloseRightBox)

        # SHOW APP
        # ///////////////////////////////////////////////////////////////
        self.show()

        # SET CUSTOM THEME
        # ///////////////////////////////////////////////////////////////
        useCustomTheme = False
        themeFile = "themes\py_dracula_light.qss"

        self.array_keys = [4]
        # SET THEME AND HACKS
        if useCustomTheme:
            # LOAD AND APPLY STYLE
            UIFunctions.theme(self, themeFile, True)

            # SET HACKS
            AppFunctions.setThemeHack(self)

        # SET HOME PAGE AND SELECT MENU
        # ///////////////////////////////////////////////////////////////

        widgets.stackedWidget.setCurrentWidget(widgets.home)
        widgets.btn_home.setStyleSheet(UIFunctions.selectMenu(widgets.btn_home.styleSheet()))
        self.option = "global"
        data = [70, 35, 40, 40]
        # spacer = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)
        widgets.create_3d_pie_chart(data, self.colors_chart, self.gl_view)
        qw = QWidget(widgets.new_page)
        ql = QVBoxLayout(qw)

        ql.addWidget(legend_list, alignment=Qt.AlignTop | Qt.AlignRight)
        # ql.addWidget(self.gl_view)
        # ql.addWidget(legend_list, alignment=Qt.AlignBottom | Qt.AlignRight)
        # widgets.layout_splitter.addWidget(self.gl_view)
        qw.setLayout(ql)

        widgets.layout_splitter.addWidget(self.gl_view, stretch=5)
        widgets.layout_splitter.addWidget(qw, stretch=1)


        # Set the legend list to align top-right

        # widgets.splitter.addWidget(qw)
        # self.LoadExcel(widgets.tableWidget)


        # self.TableSorting(widgets.tableWidget, 'DESC', 3, "global")


    # BUTTONS CLICK
    # Post here your functions for clicked buttons
    # ///////////////////////////////////////////////////////////////
    # def mousePressEvent(self, event):
    #     print('fewfewf')
    #     # Get the global mouse position

    def sorting_design(self, button):
            btnName = button.objectName()
            num = 0
            if "Asc" in btnName:
                num = int(btnName[-1]) * 2 - 1
            elif "Desc" in btnName:
                num = int(btnName[-1]) * 2 - 2
            self.buttonarray[num].setStyleSheet("""
                    """)
            button.setStyleSheet("background-color: #ff79c6;")

    def buttonClick(self):
        # GET BUTTON CLICKED
        btn = self.sender()
        btnName = btn.objectName()

        # SHOW HOME PAGE
        if btnName == "btn_home":
            widgets.stackedWidget.setCurrentWidget(widgets.home)
            UIFunctions.resetStyle(self, btnName)
            btn.setStyleSheet(UIFunctions.selectMenu(btn.styleSheet()))
            self.current_page = "home"
            self.Must_have()

        # SHOW WIDGETS PAGE
        if btnName == "btn_widgets":
            self.reset_selected_factory()
            widgets.stackedWidget.setCurrentWidget(widgets.widgets)
            UIFunctions.resetStyle(self, btnName)
            btn.setStyleSheet(UIFunctions.selectMenu(btn.styleSheet()))
            if self.dictionary == {} or self.firstTable==False:
                self.LoadExcel(widgets.tableWidget)
                self.firstTable = True
            self.current_page = "widgets"
            self.Must_have()

        # SHOW NEW PAGE
        if btnName == "btn_new":
            self.reset_selected_factory()
            widgets.stackedWidget.setCurrentWidget(widgets.new_page) # SET PAGE
            UIFunctions.resetStyle(self, btnName) # RESET ANOTHERS BUTTONS SELECTED
            btn.setStyleSheet(UIFunctions.selectMenu(btn.styleSheet())) # SELECT MENU
            # if self.dictionary.keys() == {}:
            # if self.secondTable == False:
            #     self.LoadExcel(widgets.tableWidgetSecond)
            #     self.secondTable=True
            self.current_page = "new"
            self.Must_have()
        if btnName == "loginButton":
            username = widgets.username_input.text()
            password = widgets.password_input.text()
            if username in self.accounts and password == self.accounts[username]:
                self.is_authorized = True
                widgets.btn_widgets.setVisible(True)
                widgets.btn_new.setVisible(True)
                widgets.login_button.setVisible(False)
                widgets.password_input.setVisible(False)
                widgets.username_input.setVisible(False)
                widgets.labelLogin.setVisible(False)
                widgets.labelPassword.setVisible(False)
                widgets.label_auth.setText(f"Authorization completed!\nWelcome back, {username}!")

            else:
                widgets.username_input.setText("INVALID DATA!")
        if btnName == "btn_save":
            print("Save BTN clicked!")
        if btnName == "pushButton1":
            self.OpenExcelFile()
        if btnName == "btn_export":
            self.Export_Rows()
        if btnName == "btn_select":
            selection_range = widgets.tableWidget.selectAll()
        if btnName in ["AscButton1","AscButton2","AscButton3","AscButton4", "AscButton5"] :
            self.sorting_design(btn)
            column = int(btnName[-1])
            if column not in self.array_keys:
                self.TableSorting(widgets.tableWidget, 'Asc', column)
            else:
                self.TableSorting(widgets.tableWidget, 'Asc', column)
                self.reset_button_style(btn)

        if btnName in ["DescButton1", "DescButton2", "DescButton3", "DescButton4", "DescButton5"]:
            self.sorting_design(btn)
            column = int(btnName[-1])
            if column*-1 not in self.array_keys:
                self.TableSorting(widgets.tableWidget, 'Desc', column)
            else:
                self.TableSorting(widgets.tableWidget, 'Desc', column)
                self.reset_button_style(btn)
        if btnName == "btn_reset":
            self.array_keys = [-4]
            self.TableSorting(widgets.tableWidget, 'Desc', 4)
            self.reset_button_style()

        if btnName == "btn_global":

            self.reset_selected_factory()
            self.option = "global"
            if self.current_page == 'widgets':
                self.array_keys=[]
                self.LoadExcel(widgets.tableWidget)
            elif self.current_page == 'new':
                self.gl_view.clear()
                widgets.create_3d_pie_chart(self.data_costs[self.option], self.colors_chart, self.gl_view)
            style = btn.styleSheet()
            btn.setStyleSheet(style + "\n" + "background-color: rgb(29, 34, 38)")

            # self.TableSorting(widgets.tableWidget, 'DESC', 3)


        # PRINT BTN NAME
        print(f'Button "{btnName}" pressed!')
    def reset_button_style(self, button = None):
        if button == None:
            for btn in self.buttonarray:
                btn.setStyleSheet("""""")
        else:
            button.setStyleSheet("""""")
    def Must_have(self):
        extraTopMenuLayout = self.ui.extraTopMenu.layout()
        for i in range(extraTopMenuLayout.count()):
            widget = extraTopMenuLayout.itemAt(i).widget()
            # print(widget.objectName() in self.set_factorys)
            if widget.objectName() == 'btn_share':
                widget.setVisible(self.current_page == 'home' )
            elif widget.objectName() == 'btn_more':
                widget.setVisible(self.current_page == 'home' )
            elif widget.objectName() == 'btn_adjustments':
                widget.setVisible(self.current_page == 'home' )
            elif widget.objectName() in list(self.set_factorys) or widget.objectName() == "btn_global":
                widget.setVisible( self.current_page == 'widgets' or self.current_page == 'new')

        extrapMenuLayout = self.ui.topMenus.layout()
        for i in range(extrapMenuLayout.count()):
            widget = extrapMenuLayout.itemAt(i).widget()
            if widget.objectName() == 'btn_export':
                widget.setVisible(self.current_page == 'widgets')

    def openCloseLeftBox(self, page):
            UIFunctions.toggleLeftBox(self, True, page)
            print(page)

    def search_table(self, search_text):
        if search_text in self.dictionary or search_text.lower() in self.dictionary:
                # show only the row that matches the user input key
                items = widgets.tableWidget.findItems(search_text, QtCore.Qt.MatchExactly)
                row = items[0].row()
                print(row)
                widgets.tableWidget.setCurrentItem(widgets.tableWidget.item(row, 0))
        else:
            widgets.tableWidget.setCurrentItem(widgets.tableWidget.item(0, 0))

    def OpenExcelFile(self):

        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.ExistingFile)
        dialog.setNameFilter("Excel Files (*.xlsx *.xls)")
        if dialog.exec() == QDialog.Accepted:
            self.tempaddr = dialog.selectedFiles()[0]
            if self.tempaddr[-3:] == "xls":
                x2x = XLS2XLSX(self.tempaddr)
                x2x.to_xlsx(self.tempaddr + 'x')
                self.tempaddr += 'x'

            self.flag_generate_factorys = True
            self.secondTable=False

            self.data["file"] = self.tempaddr
            self.dictionary_all_values = dict()
            for i in list(sorted(self.factory_keys)):
                widget1 = self.findChildren(QWidget, i)
                for j in widget1:
                    widgets.verticalLayout_11.removeWidget(j)
                    j.deleteLater()


            self.dict_values_factory = []
            self.buttons_red = {}
            self.buttons_orange = {}
            self.buttons_yellow = {}
            self.full_costs = {"global":0 }
            self.data_costs = {"global":[0,0,0,0,0]}
            self.set_factorys = set()
            self.array_keys = [4]
            self.another_dict = {}
            self.dictionary_all_values = {}
            self.LoadExcel(widgets.tableWidget)
        with open("variables.json", "w") as outfile:
            outfile.write(json.dumps(self.data, indent=4))
    def LoadExcel(self, widget):

        for row in range(widget.rowCount()):

            for column in range(widget.columnCount()):
                item = widget.item(row, column)
                if item is not None and row!=0:
                    item.setText("")
                    widget.setRowHidden(row, False)
        self.factory_keys = set()
        if len(self.tempaddr) > 0:
                arr_of_sheets = (openpyxl.load_workbook(self.tempaddr, read_only=True)).sheetnames
                Table = openpyxl.load_workbook(self.tempaddr)
                Sheet = Table[arr_of_sheets[0]]
                a = 0
                for i in Sheet.iter_rows():
                    if (a != 0):
                        self.factory_keys.add(str(i[5].value))
                        self.set_factorys.add(str(i[5].value))
                    a += 1

                for i in self.factory_keys:
                    self.another_dict[str(i)] = []
                    self.forecolors[str(i)] = []
                    self.data_costs[str(i)] = [0,0,0,0,0]
                    self.full_costs[str(i)] = 0
                a=0

                for i in Sheet.iter_rows():
                    if a == 0:
                        print(str(i[0].value))
                    if a == 1:
                        print(str(i[0].value) + " " + str(i[1].value))
                    if(a!=0):

                        number = str(i[self.a_number_column].value)
                        descr = str(i[self.a_descr_column].value)
                        havetb = int(i[self.havetb_column].value)
                        stock = int(i[self.stock_column].value)
                        price = float(i[self.price_column].value)
                        factory = str(i[self.factory_column].value)
                        diff_num = havetb - stock
                        if diff_num> 0:
                            self.full_costs[factory] += round(diff_num * price,1)
                            self.full_costs["global"] += round(diff_num * price, 1)
                        else:
                            self.full_costs[factory] += round(diff_num * price*(-1), 1)
                            self.full_costs["global"] += round(diff_num * price*(-1), 1)

                        match diff_num:
                            case diff_num if diff_num >= 100:
                                self.data_costs[factory][0] += round(price * diff_num, 1)
                                self.data_costs["global"][0] += round(price * diff_num, 1)
                            case diff_num if 50 <= diff_num < 100:
                                self.data_costs[factory][1] += round(price * diff_num, 1)
                                self.data_costs["global"][1] += round(price * diff_num, 1)
                            case diff_num if 0 < diff_num < 50:
                                self.data_costs[factory][2] += round(price * diff_num, 1)
                                self.data_costs["global"][2] += round(price * diff_num, 1)
                            case diff_num if (-1 * havetb) <= diff_num <= 0:
                                self.data_costs[factory][3] += round(price * diff_num * -1, 1)
                                self.data_costs["global"][3] += round(price * diff_num * -1, 1)
                            case _:
                                self.data_costs[factory][4] += round(price * diff_num * -1, 1)
                                self.data_costs["global"][4] += round(price * diff_num * -1, 1)

                        self.keys.append(number)
                        self.values.append(number+" " +descr + " " + factory)

                        diff_up_to_10 = int(round((diff_num)/10 + (0.5 if diff_num>=0 else -0.5)))*10
                        all_values = [number, havetb, stock,diff_num, diff_up_to_10, factory, price]

                        self.keys_for_dict.append(str(i[0].value))
                        self.dict_values_factory.append(all_values)
                        self.another_dict[factory].append(all_values)
                        if a==1:
                            print(self.keys_for_dict)

                    a+=1
                # print(self.data_costs['FACTORY 10'])
                self.dictionary = dict(zip(self.keys,self.values))
                self.dictionary_all_values = dict(zip(self.keys_for_dict, self.dict_values_factory))
                # print(self.dictionary_all_values)

                if self.flag_generate_factorys == True:
                    sorted_keys = sorted(self.factory_keys)
                    for i in range(0,len(sorted_keys)):
                            name = sorted_keys.pop(0)
                            button = QPushButton(f"{name}")
                            button.setObjectName(f"{name}")
                            button.setMinimumSize(QSize(0, 45))
                            button.setCursor(QCursor(Qt.PointingHandCursor))
                            button.setLayoutDirection(Qt.LeftToRight)
                            button.setStyleSheet(
                                u"background-image: url(images/icons/cil-share-boxed.png);")
                            button.clicked.connect(partial( self.ForButtons,button, widget))
                            widgets.verticalLayout_11.addWidget(button)
                    self.flag_generate_factorys = False
                    self.generate_table_factorys()
        for i in range(len(self.data_costs[self.option])):
           self.data_costs[self.option][i] = round(float(self.data_costs[self.option][i]), 1)#100 * (self.data_costs[self.option][i] / self.full_costs[self.option]), 1)


        self.TableSorting(widgets.tableWidget, 'Des', 4)
        # widget.sortItems(3, QtCore.Qt.AscendingOrder, 2, 5)

    def ForButtons(self, button,widget):
        self.array_keys = []
        self.option = button.objectName()
        if self.current_page == 'widgets':
            self.LoadExcel(widget)
        elif self.current_page == 'new':
            self.gl_view.clear()
            widgets.create_3d_pie_chart(self.data_costs[self.option], self.colors_chart, self.gl_view)
        self.Must_have()
        self.reset_selected_factory()
        widgets.btn_global.setStyleSheet((widgets.btn_global.styleSheet()).replace("background-color: rgb(29, 34, 38)", "") )
        self.sorting_design(widgets.button8)
        style = button.styleSheet()
        button.setStyleSheet(style+ "\n"+"background-color: rgb(29, 34, 38)")


    def reset_selected_factory(self):
        for i in self.set_factorys:
            my_button = self.findChild(QtWidgets.QPushButton, i)
            my_button.setStyleSheet(u"background-image: url(images/icons/cil-share-boxed.png);")
    def TableSorting(self, widget, method, sort_column):
        widget.setRowCount(1)
        widget.setRowCount(len(self.dictionary_all_values)+1)
        row_data_list = []
        if method == 'Asc':
            sort_column = sort_column
        else:
            sort_column= sort_column*-1

        if sort_column in self.array_keys:
            self.array_keys.remove(sort_column)

        elif sort_column * (-1) in self.array_keys:
            self.array_keys[self.array_keys.index(sort_column * (-1))] = sort_column
        elif sort_column not in self.array_keys:
            self.array_keys.append(sort_column)

        if self.option == "global":
            for row in range(0, len(self.dictionary_all_values)):
                row_data = []
                for col in range(widget.columnCount()):
                        item = list(self.dictionary_all_values.values())[row]

                        if item != None and col != 0:
                            row_data.append(int(item[col]))
                        if item != None and col == 0:
                            row_data.append((item[col]))
                if row_data != []:
                    row_data_list.append(row_data)

        else:
            for row in range(0, len(self.another_dict[(self.option)])):
                row_data = []
                for col in range(widget.columnCount()):
                        item = self.another_dict[(self.option)][row]
                        if item != None and col != 0 :
                            row_data.append(item[col])
                        if item != None and col == 0:
                            row_data.append(item[col])
                if row_data != []:
                    row_data_list.append(row_data)


        row_data_list.sort(key=lambda x: tuple(x[col - 1] if col > 0  else -x[(col*-1) - 1] if col != -1 else int(''.join(filter(str.isdigit, str(x[col*-1 - 1])))) for col in self.array_keys), reverse=True if col==-1  else False)

        for row, row_data in enumerate(row_data_list, 1):
            for col, cell_data in enumerate(row_data):
                item = QTableWidgetItem()
                item.setData(QtCore.Qt.DisplayRole, cell_data)
                item.setForeground(self.colortype_func(row_data[1], row_data[2]))
                widget.setItem(row, col, item)

        for row in range(widget.rowCount()):
            if row != 0:
                empty = True
                for column in range(widget.columnCount()):
                    item = widget.item(row, column)
                    if item is not None and not item.text().strip() == '':
                        empty = False
                        self.end_non_empty = row
                        break

                widget.setRowHidden(row, empty)
    def generate_table_factorys(self):
        widgets.table_factories.setRowCount(1)
        widgets.table_factories.setRowCount(len(self.another_dict.keys()) + 2)



        for i in range(1, len(self.another_dict.keys())+1):
            self.buttons_red[list(self.another_dict.keys())[i-1]] = False
            self.buttons_orange[list(self.another_dict.keys())[i - 1]] = False
            self.buttons_yellow[list(self.another_dict.keys())[i - 1]] = False
            item = QTableWidgetItem()
            item.setData(QtCore.Qt.DisplayRole, list(self.another_dict.keys())[i-1])
            item3 = QTableWidgetItem()
            item2 = QWidget()
            layout = QHBoxLayout(item2)
            button1 = QPushButton()
            button1.setObjectName(f"RedFactory{i}")
            # Set the button's style sheet
            button1.setStyleSheet("background-color: red; border-radius: 25px; text-align: center;")
            button1.setFixedSize(20, 20)
            button1.clicked.connect(partial(self.total_cost_color,"Red", i, button1))
            layout.addWidget(button1)
            button2 = QPushButton()
            button2.setObjectName(f"OrangeFactory{i}")
            button2.clicked.connect(partial(self.total_cost_color,"Orange", i, button2))
            # Set the button's style sheet
            button2.setStyleSheet("background-color: orange; border-radius: 25px; text-align: center;")
            button2.setFixedSize(20, 20)
            layout.addWidget(button2)
            button3 = QPushButton()
            button3.setObjectName(f"YellowFactory{i}")
            button3.clicked.connect(partial(self.total_cost_color, "Yellow", i, button3))
            # Set the button's style sheet
            button3.setStyleSheet("background-color: yellow; border-radius: 25px; text-align: center;")
            button3.setFixedSize(20, 20)
            layout.addWidget(button3)
            widgets.table_factories.setCellWidget(i, 2, item2)
            widgets.table_factories.setItem(i, 0, item)
            widgets.table_factories.setItem(i, 1, item3)

    def total_cost_color(self, color, row_table, button):

            factory_value_row = list(self.another_dict.keys())[row_table-1]
            if color == "Red":
                max_value = (10000)
                min_value =100
                self.buttons_red[factory_value_row] = not self.buttons_red[factory_value_row]
                button_clicked_change = self.buttons_red[factory_value_row]


            elif color == "Orange":
                min_value = 50
                max_value = 100
                self.buttons_orange[factory_value_row] = not self.buttons_orange[factory_value_row]
                button_clicked_change = self.buttons_orange[factory_value_row]

            else:
                min_value = 0
                max_value=50
                self.buttons_yellow[factory_value_row] = not self.buttons_yellow[factory_value_row]
                button_clicked_change = self.buttons_yellow[factory_value_row]

            self.border_color_clicked(button, button_clicked_change)
    # print(self.another_dict[str(list(self.another_dict)[row_table])])

            if color in self.forecolors[factory_value_row]:
                self.forecolors[factory_value_row].remove(color)
            else:
                self.forecolors[factory_value_row].append(color)
            if len(self.forecolors[factory_value_row]) == 0:
                forecolor = '#000000'
            elif len(self.forecolors[factory_value_row]) == 1:
                forecolor = self.forecolors[factory_value_row][0]
            elif len(self.forecolors[factory_value_row]) == 2:
                forecolor = '#0362fc'
            else:
                forecolor = '#FFFFFF'
            print(self.forecolors[factory_value_row])
            print(forecolor)
            sum_row = widgets.table_factories.item(row_table, 1)
            factory_row = widgets.table_factories.item(row_table, 0)

            if sum_row and sum_row.text()!='':
                sum = float(sum_row.text())
            else:
                sum = 0
            # sum = int(sum_row.text()) if sum_row.text() is not None else 0
            if factory_row != None:
                for properties in self.another_dict[factory_row.text()]:
                    # print(properties)
                    diff = (properties[1]-properties[2])
                    if min_value <= diff < max_value:
                        if button_clicked_change:
                            sum += round(properties[6] * diff, 1)
                        else:
                            sum -= round(properties[6] * diff, 1)
                print(sum)
                sum = round(sum,1)
                if sum == 0:
                    sum =''
                sum_row.setData(QtCore.Qt.DisplayRole, sum)
                sum_row.setForeground(QColor(forecolor))
                # print("Final summ " + str(sum))
    def border_color_clicked(self, button, status):
        if status:
            style = button.styleSheet()
            button.setStyleSheet(style + "border: 2px solid magenta")
        else:
            style = button.styleSheet()
            style = style.replace("border: 2px solid magenta", "")
            button.setStyleSheet(style)


    def colortype_func(self, havetb, stock):
        diff = havetb - stock
        match diff:
            case diff if diff >= 100:
                color = QColor(255, 0, 0)
            case diff if 50 <= diff < 100:
                color = QColor(255, 127, 0)
            case diff if 0 < diff < 50:
                color = QColor(255, 255, 0)
            case diff if (-1 * havetb) <= diff <= 0:
                color = QColor(0, 255, 0)
            case _:
                color = QColor(255, 255, 255)
        return color
    def Export_Rows(self):
        selectionModel = widgets.tableWidget.selectionModel()
        selectedRanges = selectionModel.selectedRows()

        # Create a new workbook
        wb = openpyxl.Workbook()

        # Add a new worksheet
        ws = wb.active
        h = ["Number", "Have to be", "Now we have", "Diff", "Diff up to 10"]
        # ws.append(h)

        # Create a progress dialog
        progressDialog = QProgressDialog("Exporting rows...", "Cancel", 0, len(selectedRanges), self)
        progressDialog.setWindowModality(Qt.WindowModal)
        progressDialog.setValue(0)
        progressDialog.setWindowTitle("Export")
        progressDialog.setCancelButton(None)
        # Export the selected rows to Excel
        realrow = 1
        for i in range (5):
            ws.cell(row=realrow, column=i + 1).fill =PatternFill(patternType='solid', fgColor=Color("538DD5"))
            ws.cell(row=realrow, column=i+1).value = h[i]
        realrow+=1
        for i, r in enumerate(selectedRanges):
            # Update the progress dialog
            progressDialog.setValue(i)
            if progressDialog.wasCanceled():
                break

            # Get the items in the row
            rowItems = [widgets.tableWidget.item(r.row(), c) for c in
                        range(widgets.tableWidget.columnCount())]

            rowValues = []
            c=0
            for item in rowItems:
                if item is None:
                    break
                cellValue = item.text()
                cellForeground = item.foreground()
                # if cellValue =='Number':
                #     break

                # Get the color components
                red = hex(cellForeground.color().red())[2:].upper().zfill(2)
                blue = hex(cellForeground.color().blue())[2:].upper().zfill(2)
                green = hex(cellForeground.color().green())[2:].upper().zfill(2)

                color = red + green + blue
                if color == "000000":
                    color = "538DD5"
                # print(color)
                fill = PatternFill(patternType='solid', fgColor=Color(color))

                ws.cell(row=realrow, column=c + 1).fill = fill
                ws.cell(row=realrow, column=c + 1).value = cellValue
                rowValues.append(cellValue)
                c+=1
            if r.row() == 1:
                continue
            # ws.append(rowValues)
            realrow += 1

        # Close the progress dialog
        progressDialog.setValue(len(selectedRanges))

        # Save the workbook
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getSaveFileName(self, "Save Excel file", "", "Excel Files (*.xlsx)", options=options)
        if fileName:
            wb.save(fileName)
    def CellClicked(self, row, column):
        item = widgets.tableWidget.item(row, 0)
        if item is not None and row != 0:
            widgets.plainTextEdit.setPlainText(str(self.dictionary[item.text()]))

    def resizeEvent(self, event):
        # Update Size Grips
        UIFunctions.resize_grips(self)

    # MOUSE CLICK EVENTS
    # ///////////////////////////////////////////////////////////////
    def mousePressEvent(self, event):
        # SET DRAG POS WINDOW
        self.dragPos = event.globalPos()

        # PRINT MOUSE EVENTS
        if event.buttons() == Qt.LeftButton:
            print('Mouse click: LEFT CLICK')
            start_point = event.pos()

            # Create the end point 20 pixels up
            end_point = QPoint(start_point.x(), start_point.y() - 20)

            # Create a line between the start and end points
            line = QLine(start_point, end_point)

            # Add the line to the list of lines to be drawn
            self.lines.append(line)
            # Draw all lines
        self.update()
        if event.buttons() == Qt.RightButton:
            print('Mouse click: RIGHT CLICK')
        globalPos = QCursor.pos()
        self.mouse_position = event.pos()


        # Create a QPixmap, and render the screen content into it
        screen = QApplication.primaryScreen()
        print(self.height())
        if screen is not None:
            pixmap = screen.grabWindow(0, globalPos.x(), globalPos.y(), 1, 1)
        else:
            return

        # Get the color of the pixel under the mouse cursor
        color = QColor(pixmap.toImage().pixelColor(0, 0)).getRgbF()[:3]

        if color[0] == color[1] == color[2]:  # white
            if self.is_white(color):
                self.labelHello.setText(f"{round(100*self.data_costs[self.option][4]/self.full_costs[self.option], 1)}% \u20AC{round(self.data_costs[self.option][4],1)} "  )
                self.labelHello.move(QPoint(event.pos().x()- 150, self.height() / 8 * 4.6))
                self.labelHello.setStyleSheet("color: white; font-family: Arial; font-size: 16px; border: 0.5px solid white; padding-top: 5px; padding-bottom: 5px; padding-left: 10px; padding-right: 10px;")

        elif color[0] == color[1] != 0 and color[2] == 0:  # yellow
            if self.is_yellow(color):
                self.labelHello.setText(f"{round(100*self.data_costs[self.option][2]/self.full_costs[self.option],1)}% \u20AC{round(self.data_costs[self.option][2],1)}" )
                self.labelHello.move(QPoint(event.pos().x()- 150, self.height()/8 * 1.8))
                self.labelHello.setStyleSheet(
                    "color: yellow; font-family: Arial; font-size: 16px; border: 0.5px solid yellow; padding-top: 5px; padding-bottom: 5px; padding-left: 10px; padding-right: 10px;")

        elif color[1] == color[2] == 0 and color[0] != 0:  # red
            if self.is_red(color):
                self.labelHello.setText(
                    f"{round(100 * self.data_costs[self.option][0] / self.full_costs[self.option],1)}% \u20AC{round(self.data_costs[self.option][0], 1)}")
                self.labelHello.move(QPoint(event.pos().x() - 150, self.height()/8 * 4.6))
                self.labelHello.setStyleSheet(
                    "color: #fa020f; font-family: Arial; font-size: 16px; border: 0.5px solid red; padding-top: 5px; padding-bottom: 5px; padding-left: 10px; padding-right: 10px;")


        elif color[1] != 0 and color[0] == color[2] == 0:  # green
            if self.is_green(color):
                self.labelHello.setText(
                    f"{round(100 * self.data_costs[self.option][3] / self.full_costs[self.option],1)}% \u20AC{round(self.data_costs[self.option][3], 1)}")
                self.labelHello.move(QPoint(event.pos().x()- 150, self.height()/8 * 1.8 ))
                self.labelHello.setStyleSheet(
                    "color: #15b02f;font-family: Arial; font-size: 16px; border: 0.5px solid green; padding-top: 5px; padding-bottom: 5px; padding-left: 10px; padding-right: 10px;")


        elif color[1] != color[0] and color[1] != 0 and 1.52 <= round(color[0] / color[1], 2) <= 1.56:
            if self.is_orange(color):
                self.labelHello.setText(
                    f"{round(100 * self.data_costs[self.option][1] / self.full_costs[self.option],1)}% \u20AC{round(self.data_costs[self.option][1], 1)}")

                self.labelHello.move(QPoint(event.pos().x() - 150, self.height()/8 * 4.6))
                self.labelHello.setStyleSheet(
                    "color: orange; font-family: Arial; font-size: 16px; border: 0.5px solid orange; padding-top: 5px; padding-bottom: 5px; padding-left: 10px; padding-right: 10px;")

        else:
            self.labelHello.setText('')
            self.labelHello.setStyleSheet("border: none")



        super().mouseMoveEvent(event)

    def paintEvent(self, event):
        painter = QPainter(self)

        # Draw all lines
        for line in self.lines:
            painter.setPen(self.pen)
            painter.drawLine(line)

    def is_orange(self, color):
        return all(self.min_orange[i] <= color[i] <= self.max_orange[i] for i in range(3))

    def is_yellow(self, color):
        return all(self.min_yellow[i] <= color[i] <= self.max_yellow[i] for i in range(3))

    def is_red(self, color):
        return all(self.min_red[i] <= color[i] <= self.max_red[i] for i in range(3))

    def is_green(self, color):
        return all(self.min_green[i] <= color[i] <= self.max_green[i] for i in range(3))

    def is_white(self, color):
        return all(self.min_white[i] <= color[i] <= self.max_white[i] for i in range(3))

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon("icon.ico"))
    window = MainWindow()
    # window.setMouseTracking(True)
    sys.exit(app.exec())
